import openpyxl

def update_scores():
    # 打开 Excel 文件
    wb = openpyxl.load_workbook('D:\\ddd\\FDU\\助教\\项目管理\\打分\\2023项目管理课程分数明细表1.xlsx', data_only=True)

    # 获取“组内评分统计”工作表
    sheet1 = wb['平时分和论文']
    
    # 获取“期末论文分数”工作表
    sheet2 = wb['总分']
    
    # 遍历 sheet1 工作表的每一行（从第二行开始）
    for row in sheet1.iter_rows(min_row=3, values_only=True):
        score = row[14]  # 获取分数列
        print(score)

        if score is None:
            break

        # for i in range(4):
        #     name = row[i + 1]
        name = row[1]
        print(name, score)

        # 在“期末论文分数”工作表中查找名字相同的行
        if name is not None and score is not None:
            for row_number, row2 in enumerate(sheet2.iter_rows(min_row=2, values_only=True)):
                if row2[1] == name:  # 找到匹配的名字
                    # 更新相应行的E列分数
                    # print(row2[0])
                    sheet2.cell(row=row_number + 2, column=3).value = score
                    break
    
    # 保存修改后的 Excel 文件
    wb.save('D:\\ddd\\FDU\\助教\\项目管理\\打分\\2023项目管理课程分数明细表1.xlsx')


def register_scores():
    # 打开 Excel 文件
    wb = openpyxl.load_workbook('D:\\ddd\\FDU\\助教\\项目管理\\打分\\2023项目管理课程分数明细表1.xlsx', data_only=True)
    wb2 = openpyxl.load_workbook('D:\\ddd\\FDU\\助教\\项目管理\\打分\\项目管理成绩.xlsx')

    # 获取“组内评分统计”工作表
    sheet1 = wb['总分']
    
    # 获取“期末论文分数”工作表
    sheet2 = wb2['sheet1']
    
    # 遍历 sheet1 工作表的每一行（从第二行开始）
    for row in sheet1.iter_rows(min_row=2, values_only=True):
        score = row[6]  # 获取分数列
        name = row[1]

        if score is None or name is None:
            break
        
        print(name, score)

        for row_number, row2 in enumerate(sheet2.iter_rows(min_row=9, values_only=True)):
            if row2[2] == name:  # 找到匹配的名字
                # 更新相应行的分数
                sheet2.cell(row=row_number+9, column=5).value = score
                break

    
    # 保存修改后的 Excel 文件
    wb2.save('D:\\ddd\\FDU\\助教\\项目管理\\打分\\bbb.xlsx')
# 调用函数更新分数
# update_scores()
register_scores()
